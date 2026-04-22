"""
OEE ETL Pipeline
================
Reads daily IML OEE Excel files and produces a single JSON file
that the OEE dashboard (oee_dashboard_v2.html) reads directly.

REQUIREMENTS
------------
Install once, in Command Prompt:
    py -m pip install openpyxl pandas

USAGE
-----
Place all your daily Excel files in one folder, then run:

    py oee_etl.py

By default it looks for Excel files in the same folder as this script
and writes oee_data.json next to it.

To specify a different folder:
    folder = "data"

To write the JSON somewhere else:
    py oee_etl.py --output "D:\WORKING\OEE\\output\\oee_data.json"

FILE NAMING
-----------
Each Excel file must contain the date somewhere in its name as DD_MM_YYYY.
Examples that all work:
    21_03_2026.xlsx
    OEE_21_03_2026.xlsx
    1774355103920_21_03_2026.xlsx

FORMULA METHODOLOGY
--------------------
All metrics follow the Sheet1 methodology from your Excel workbook exactly:

  Shift run %       = machines that ran / total machine-slots  (from Shift tracker)
  Availability loss = (sum shift_hrs - sum run_hrs) / sum shift_hrs
  Cavity loss       = (sum total_cav - sum running_cav) / sum total_cav
  Quality loss      = sum rejected_pcs / (sum rejected + sum good)
  OEE (running)     = sum actual_good / sum target             (Sheet1 R6)
  Speed loss        = residual: OEE_running / (avail x cav x quality)
  Overall OEE       = OEE_running x shift_run_pct              (Sheet1 D10)

  Downtime %        = each category hrs / total shift hrs      (Sheet2 methodology)
  Not-run reasons   = count per reason / total machine-slots
"""

import os
import re
import sys
import json
import argparse
import logging
from datetime import datetime
from pathlib import Path

# ── Dependency check (friendly error before import fails) ────────────────────
missing = []
try:
    import pandas as pd
except ImportError:
    missing.append("pandas")
try:
    from openpyxl import load_workbook
except ImportError:
    missing.append("openpyxl")

if missing:
    print("\nERROR: Missing required libraries. Please run this in Command Prompt:")
    print(f"\n    py -m pip install {' '.join(missing)}\n")
    sys.exit(1)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("oee_etl")

# ── Column positions in OEE tracker (1-indexed, for openpyxl) ────────────────
# Two layouts supported — detected automatically from row 10 header:
#
#   OLD layout (col I = Product type):   no Operator name column
#   NEW layout (col I = Operator name):  inserted at col 9, all others +1
#
# Detection: if cell (row=10, col=9) contains "Operator", use NEW layout.

OEE_COL_OLD = {
    "date":             2,   # B
    "shift":            3,   # C
    "duration_hrs":     6,   # F
    "machine_no":       7,   # G
    "to_fill":          8,   # H
    "product_type":     9,   # I
    "product_name":     10,  # J
    "mould_no":         11,  # K
    "weight_gm":        13,  # M
    "rated_cycle_sec":  14,  # N
    "total_cavities":   15,  # O
    "target_pcs":       16,  # P
    "actual_good_pcs":  17,  # Q
    "rej_label_kg":     19,  # S
    "rej_plain_kg":     20,  # T
    "rej_trial_kg":     21,  # U
    "total_rej_pcs":    22,  # V
    "run_hrs":          26,  # Z
    "dt_label":         27,  # AA
    "dt_colour":        28,  # AB
    "dt_mould":         29,  # AC
    "dt_label_unavail": 30,  # AD
    "dt_proc_fail":     31,  # AE
    "dt_mach_bkdn":     32,  # AF
    "dt_mould_bkdn":    33,  # AG
    "dt_other":         34,  # AH
    "avail_loss_pct":   36,  # AJ
    "running_cavities": 37,  # AK
    "perf_cav_loss":    38,  # AL
    "actual_cycle_sec": 39,  # AM
    "perf_spd_loss":    40,  # AN
    "overall_perf_loss":41,  # AO
    "oee_pct":          42,  # AP
}

OEE_COL_NEW = {
    "date":             2,   # B
    "shift":            3,   # C
    "duration_hrs":     6,   # F
    "machine_no":       7,   # G
    "to_fill":          8,   # H
    "operator_name":    9,   # I  ← new column
    "product_type":     10,  # J
    "product_name":     11,  # K
    "mould_no":         12,  # L
    "weight_gm":        14,  # N
    "rated_cycle_sec":  15,  # O
    "total_cavities":   16,  # P
    "target_pcs":       17,  # Q
    "actual_good_pcs":  18,  # R
    "rej_label_kg":     20,  # T
    "rej_plain_kg":     21,  # U
    "rej_trial_kg":     22,  # V
    "total_rej_pcs":    23,  # W
    "run_hrs":          27,  # AA
    "dt_label":         28,  # AB
    "dt_colour":        29,  # AC
    "dt_mould":         30,  # AD
    "dt_label_unavail": 31,  # AE
    "dt_proc_fail":     32,  # AF
    "dt_mach_bkdn":     33,  # AG
    "dt_mould_bkdn":    34,  # AH
    "dt_other":         35,  # AI
    "avail_loss_pct":   37,  # AK
    "running_cavities": 38,  # AL
    "perf_cav_loss":    39,  # AM
    "actual_cycle_sec": 40,  # AN
    "perf_spd_loss":    41,  # AO
    "overall_perf_loss":42,  # AP
    "oee_pct":          43,  # AQ
}

def detect_oee_col(ws) -> dict:
    """Return the correct column map by inspecting row 10, column I (col 9)."""
    header_i = ws.cell(row=10, column=9).value
    if header_i and "operator" in str(header_i).lower():
        log.info("  Detected NEW column layout (Operator name at col I)")
        return OEE_COL_NEW
    return OEE_COL_OLD

# OEE_COL is set per-file inside parse_file() — do not use this global directly
OEE_COL = OEE_COL_OLD  # fallback default (unused after parse_file sets it)

SHIFT_COL = {
    "date":    2,  # B
    "shift":   3,  # C
    "machine": 4,  # D
    "ran":     5,  # E
    "reason":  6,  # F
}

DT_KEYS   = ["dt_label","dt_colour","dt_mould","dt_label_unavail",
             "dt_proc_fail","dt_mach_bkdn","dt_mould_bkdn","dt_other"]
DT_LABELS = ["Label change","Colour change","Mould change","Label unavail.",
             "Process failure","Mach. breakdown","Mould breakdown","Other"]


# ── Date extraction ───────────────────────────────────────────────────────────

def extract_date(path):
    """
    Find a date in the filename and return as datetime.
    Accepts any of these formats:
        DD_MM_YYYY   e.g. 21_03_2026
        DD.MM.YYYY   e.g. 21.03.2026
        DD-MM-YYYY   e.g. 21-03-2026
        DD MM YYYY   e.g. 21 03 2026
        D.MM.YYYY    e.g. 12.3.2026  (single-digit day or month)
        D.M.YYYY     e.g. 1.3.2026
    """
    stem = Path(path).stem
    # Try DD[sep]MM[sep]YYYY where sep is . _ - or space, digits can be 1 or 2
    m = re.search(r"(\d{1,2})[._\- ](\d{1,2})[._\- ](\d{4})", stem)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    log.warning(f"Could not parse date from: {Path(path).name}")
    return None


# ── Excel parsing ─────────────────────────────────────────────────────────────

def parse_file(path):
    """
    Parse one daily Excel file.
    Returns dict with keys: date_str, all_slots, not_run, records
    or None if the file cannot be parsed.
    """
    file_date = extract_date(path)
    if file_date is None:
        return None

    date_str = file_date.strftime("%Y-%m-%d")
    log.info(f"Reading: {Path(path).name}  ->  {date_str}")

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        log.error(f"  Cannot open file: {e}")
        return None

    if "Shift tracker" not in wb.sheetnames or "OEE tracker" not in wb.sheetnames:
        log.error(f"  Missing required sheets. Found: {wb.sheetnames}")
        return None

    # ── Shift tracker ──────────────────────────────────────────────────────
    ws_shift = wb["Shift tracker"]
    not_run  = []
    all_slots = []

    for row in range(15, 300):
        machine = ws_shift.cell(row=row, column=SHIFT_COL["machine"]).value
        shift   = ws_shift.cell(row=row, column=SHIFT_COL["shift"]).value
        ran     = ws_shift.cell(row=row, column=SHIFT_COL["ran"]).value

        if machine is None or shift is None:
            if row > 20:
                break
            continue

        try:
            machine_no = int(machine)
        except (ValueError, TypeError):
            continue

        shift_str  = str(shift).strip()
        ran_str    = str(ran).strip().upper() if ran is not None else "Y"
        reason_raw = ws_shift.cell(row=row, column=SHIFT_COL["reason"]).value
        reason_str = str(reason_raw).strip() if reason_raw else None

        slot = {"date": date_str, "shift": shift_str, "machine_no": machine_no}
        all_slots.append(slot)

        if ran_str == "N":
            not_run.append({
                **slot,
                "reason": reason_str or "Missing reason code",
            })

    log.info(f"  Shift tracker: {len(all_slots)} slots, {len(not_run)} not run")

    # ── OEE tracker ────────────────────────────────────────────────────────
    ws_oee  = wb["OEE tracker"]
    OEE_COL = detect_oee_col(ws_oee)  # auto-detect old vs new layout
    records = []

    # ── Find last row with data so we don't break early ───────────────────
    # Scan once to find where data ends (last row with any value in key cols)
    last_row = 11
    for _r in range(11, 2000):
        has_data = any(
            ws_oee.cell(row=_r, column=c).value is not None
            for c in [2, 7, 8, 16, 17]  # date, machine, to_fill, target, actual
        )
        if has_data:
            last_row = _r

    for row in range(11, last_row + 1):
        # ── Row inclusion logic — matches Excel's SUMIF(P>0) exactly ──────
        # Excel includes a row if target_pcs > 0, regardless of to_fill flag.
        # Some rows are marked "No" but have real production data filled in
        # (operator error or copy-paste). We must include those too.
        # We require both target_pcs > 0 AND actual_good_pcs > 0 to avoid
        # picking up formula-populated but genuinely empty rows.
        def get(col_name, row=row):
            return ws_oee.cell(row=row, column=OEE_COL[col_name]).value

        def num(col_name, default=0, row=row):
            v = ws_oee.cell(row=row, column=OEE_COL[col_name]).value
            try:
                return float(v) if v is not None else default
            except (ValueError, TypeError):
                return default

        target_val = num("target_pcs")
        actual_val = num("actual_good_pcs")

        # Skip rows with no real production data
        if target_val <= 0 or actual_val <= 0:
            continue

        machine_no = num("machine_no")
        if machine_no == 0:
            continue

        records.append({
            "date":             date_str,
            "shift":            str(get("shift") or "").strip(),
            "machine_no":       int(machine_no),
            "operator_name":    str(get("operator_name") or "").strip() if "operator_name" in OEE_COL else "",
            "product_type":     str(get("product_type") or "").strip(),
            "product_name":     str(get("product_name") or "").strip(),
            "mould_no":         str(get("mould_no") or "").strip(),
            "duration_hrs":     num("duration_hrs"),
            "run_hrs":          num("run_hrs"),
            "total_cavities":   int(num("total_cavities")),
            "running_cavities": min(int(num("running_cavities")), int(num("total_cavities"))),  # cap at total — data entry errors can give rc>tc
            "target_pcs":       num("target_pcs"),
            "actual_good_pcs":  num("actual_good_pcs"),
            "total_rej_pcs":    num("total_rej_pcs"),
            "rej_label_kg":     num("rej_label_kg"),
            "rej_plain_kg":     num("rej_plain_kg"),
            "rej_trial_kg":     num("rej_trial_kg"),
            "dt_label":         num("dt_label"),
            "dt_colour":        num("dt_colour"),
            "dt_mould":         num("dt_mould"),
            "dt_label_unavail": num("dt_label_unavail"),
            "dt_proc_fail":     num("dt_proc_fail"),
            "dt_mach_bkdn":     num("dt_mach_bkdn"),
            "dt_mould_bkdn":    num("dt_mould_bkdn"),
            "dt_other":         num("dt_other"),
        })

    log.info(f"  OEE tracker:   {len(records)} active machine-shift rows")

    # ── Rebuild not_run to match the new slot definition ─────────────────
    # not_run = slots in all_slots that have NO OEE record (target_pcs > 0)
    # This ensures the reason breakdown counts tie exactly to the
    # not-run count used for shift run % (total_slots - ran).
    # Reasons are looked up from the Shift tracker reason map built above.
    ran_set = {(r["date"], r["shift"], r["machine_no"]) for r in records}
    # Build a reason lookup from the original shift-tracker not_run list
    reason_lookup = {
        (nr["date"], nr["shift"], nr["machine_no"]): nr["reason"]
        for nr in not_run
    }
    not_run = [
        {
            "date":       s["date"],
            "shift":      s["shift"],
            "machine_no": s["machine_no"],
            "reason":     reason_lookup.get(
                              (s["date"], s["shift"], s["machine_no"]),
                              "Missing reason code"
                          ),
        }
        for s in all_slots
        if (s["date"], s["shift"], s["machine_no"]) not in ran_set
    ]
    log.info(f"  Not-run slots: {len(not_run)} (Shift tracker total - OEE ran)")

    return {
        "date_str":  date_str,
        "all_slots": all_slots,
        "not_run":   not_run,
        "records":   records,
    }


# ── Formula engine — matches Sheet1 exactly ───────────────────────────────────

def compute_metrics(records, not_run, all_slots):
    """
    Weighted-average approach: each metric is a per-row rate averaged
    by duration_hrs weight. Speed loss is derived as residual.
    """
    def s(fn):
        return sum(fn(r) for r in records)

    # ── Weighted-average approach for all primary metrics ─────────────────
    # Products are not like-for-like (big containers vs small lids/spoons
    # have very different cycle times and cavity counts). Each metric is a
    # weighted average of per-row rates, weight = duration_hrs, so every
    # machine-hour counts equally regardless of product type.
    # Speed is the RESIDUAL so avail×cav×speed×qual = oee_running always.
    # Note: residual speed slightly understates true speed loss — it absorbs
    # the positive covariance between factors (good shifts tend to be good
    # on all dimensions simultaneously).

    total_dur_w = s(lambda r: r["duration_hrs"])

    def wt(r, val):
        return val * (r["duration_hrs"] or 0)

    avail_loss = s(lambda r: wt(r,
        (r["duration_hrs"] - r["run_hrs"]) / r["duration_hrs"]
        if r["duration_hrs"] else 0)) / total_dur_w if total_dur_w else 0

    cav_loss = s(lambda r: wt(r,
        (r["total_cavities"] - r["running_cavities"]) / r["total_cavities"]
        if r["total_cavities"] else 0)) / total_dur_w if total_dur_w else 0

    quality_loss = s(lambda r: wt(r,
        r["total_rej_pcs"] / (r["total_rej_pcs"] + r["actual_good_pcs"])
        if (r["total_rej_pcs"] + r["actual_good_pcs"]) else 0)) / total_dur_w if total_dur_w else 0

    oee_running = s(lambda r: wt(r,
        r["actual_good_pcs"] / r["target_pcs"]
        if r["target_pcs"] else 0)) / total_dur_w if total_dur_w else 0

    # Raw sums for output display fields
    total_shift_hrs = s(lambda r: r["duration_hrs"])
    total_run_hrs   = s(lambda r: r["run_hrs"])
    total_target    = s(lambda r: r["target_pcs"])
    total_actual    = s(lambda r: r["actual_good_pcs"])
    total_rej       = s(lambda r: r["total_rej_pcs"])

    # D4: shift run %
    # Total slots = all machine-shifts in Shift tracker (ground truth)
    # Ran = unique (date, shift, machine) with target_pcs > 0 in OEE tracker
    # Not run = total_slots - ran  (delta)
    total_slots   = len(all_slots)
    ran_keys_set  = {(r["date"], r["shift"], r["machine_no"]) for r in records}
    ran_cnt       = len(ran_keys_set)
    not_ran_cnt   = total_slots - ran_cnt
    shift_run_pct = ran_cnt / total_slots if total_slots else 1

    # Speed loss: residual — guarantees avail × cav × speed × qual = oee_running
    avail_rate   = 1 - avail_loss
    cav_rate     = 1 - cav_loss
    quality_rate = 1 - quality_loss
    denom        = avail_rate * cav_rate * quality_rate
    speed_rate   = oee_running / denom if denom > 0 else 1.0
    speed_loss   = max(0.0, 1 - speed_rate)

    # D10: overall OEE = OEE_running x shift_run_pct
    overall_oee = oee_running * shift_run_pct

    # Downtime breakdown as % of total shift hrs (Sheet2 methodology)
    dt_breakdown_pct = {}
    dt_breakdown_hrs = {}
    for k, label in zip(DT_KEYS, DT_LABELS):
        hrs = s(lambda r, k=k: r[k])
        dt_breakdown_pct[label] = round(hrs / total_shift_hrs, 6) if total_shift_hrs else 0
        dt_breakdown_hrs[label] = round(hrs, 2)

    # Not-run reasons as % of total slots
    reason_counts = {}
    for nr in not_run:
        key = nr["reason"] or "Missing reason code"
        reason_counts[key] = reason_counts.get(key, 0) + 1
    reason_pcts = {
        k: round(v / total_slots, 6) if total_slots else 0
        for k, v in reason_counts.items()
    }

    return {
        "total_shift_hrs":  round(total_shift_hrs, 1),
        "total_run_hrs":    round(total_run_hrs, 1),
        "total_target":     int(total_target),
        "total_actual":     int(total_actual),
        "total_rej":        round(total_rej, 0),
        "total_slots":      total_slots,
        "ran_slots":        ran_cnt,
        "not_ran_slots":    not_ran_cnt,
        "shift_run_pct":    round(shift_run_pct, 6),
        "avail_loss":       round(avail_loss, 6),
        "cav_loss":         round(cav_loss, 6),
        "speed_loss":       round(speed_loss, 6),
        "quality_loss":     round(quality_loss, 6),
        "oee_running":      round(oee_running, 6),
        "overall_oee":      round(overall_oee, 6),
        "dt_breakdown_pct": dt_breakdown_pct,
        "dt_breakdown_hrs": dt_breakdown_hrs,
        "reason_pcts":      reason_pcts,
        "reason_counts":    reason_counts,
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="IML OEE ETL -- converts daily Excel files to dashboard JSON"
    )

    # ✅ FIX: default folder changed to data
    parser.add_argument(
        "--folder",
        default="data",
        help="Folder containing Excel files (default: data folder)"
    )

    # ✅ FIX: output always inside output folder
    parser.add_argument(
        "--output",
        default="output/oee_data.json",
        help="Output JSON path"
    )

    args = parser.parse_args()

    # ✅ FIX: absolute safe path
    base_dir = Path(__file__).parent
    folder = base_dir / args.folder
    output_path = base_dir / args.output

    if not folder.exists():
        log.error(f"Folder not found: {folder}")
        sys.exit(1)

    excel_files = sorted(
        list(folder.glob("*.xlsx")) + list(folder.glob("*.xls")),
        key=lambda p: extract_date(str(p)) or datetime.min
    )

    log.info(f"Found {len(excel_files)} Excel file(s) in {folder}")

    if not excel_files:
        log.error("No Excel files found in data folder ❌")
        sys.exit(1)

    # बाकी code SAME रहेगा 👇

    # ── Parse all files ────────────────────────────────────────────────────
    all_records  = []
    all_not_run  = []
    all_slots    = []
    daily_data   = []
    failed_files = []

    for fp in excel_files:
        result = parse_file(str(fp))
        if result is None:
            failed_files.append(fp.name)
            continue

        all_records.extend(result["records"])
        all_not_run.extend(result["not_run"])
        all_slots.extend(result["all_slots"])

        day_metrics = compute_metrics(
            result["records"],
            result["not_run"],
            result["all_slots"],
        )
        daily_data.append({
            "date":          result["date_str"],
            "overall_oee":   day_metrics["overall_oee"],
            "oee_running":   day_metrics["oee_running"],
            "shift_run_pct": day_metrics["shift_run_pct"],
            "avail_loss":    day_metrics["avail_loss"],
            "cav_loss":      day_metrics["cav_loss"],
            "speed_loss":    day_metrics["speed_loss"],
            "quality_loss":  day_metrics["quality_loss"],
            "total_actual":  day_metrics["total_actual"],
            "total_target":  day_metrics["total_target"],
        })

    if not all_records:
        log.error("No valid data could be read from any file.")
        sys.exit(1)

    # ── Aggregate across all dates ─────────────────────────────────────────
    log.info(f"Computing aggregate across {len(daily_data)} date(s)...")
    aggregate = compute_metrics(all_records, all_not_run, all_slots)

    # ── Machine-level aggregation ──────────────────────────────────────────
    machine_nos = sorted(set(r["machine_no"] for r in all_records))
    machines = []
    for m in machine_nos:
        m_recs  = [r for r in all_records if r["machine_no"] == m]
        m_nr    = [r for r in all_not_run  if r["machine_no"] == m]
        m_slots = [r for r in all_slots    if r["machine_no"] == m]
        mc = compute_metrics(m_recs, m_nr, m_slots)

        products = [r["product_name"] for r in m_recs if r["product_name"]]
        primary_product = max(set(products), key=products.count) if products else ""
        prod_types = [r["product_type"] for r in m_recs if r["product_type"]]
        primary_type = max(set(prod_types), key=prod_types.count) if prod_types else ""

        machines.append({
            "machine_no":       m,
            "primary_product":  primary_product,
            "product_type":     primary_type,
            "overall_oee":      mc["overall_oee"],
            "oee_running":      mc["oee_running"],
            "shift_run_pct":    mc["shift_run_pct"],
            "avail_loss":       mc["avail_loss"],
            "cav_loss":         mc["cav_loss"],
            "speed_loss":       mc["speed_loss"],
            "quality_loss":     mc["quality_loss"],
            "total_actual":     mc["total_actual"],
            "total_target":     mc["total_target"],
            "total_rej":        mc["total_rej"],
            "ran_slots":        mc["ran_slots"],
            "total_slots":      mc["total_slots"],
            "dt_breakdown_hrs": mc["dt_breakdown_hrs"],
        })

    # ── Build and write JSON ───────────────────────────────────────────────
    date_range = {
        "from": daily_data[0]["date"] if daily_data else "",
        "to":   daily_data[-1]["date"] if daily_data else "",
        "days": len(daily_data),
    }

    output = {
        "_meta": {
            "generated":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_files": len(daily_data),
            "failed_files": failed_files,
            "date_range":   date_range,
            "formula_notes": {
                "overall_oee":  "OEE_running x shift_run_pct  (Sheet1 D10)",
                "oee_running":  "sum(actual) / sum(target)    (Sheet1 R6)",
                "shift_run_pct":"ran_slots / total_slots       (Sheet1 D4)",
                "avail_loss":   "(shift_hrs - run_hrs) / shift_hrs  (Sheet1 AI6)",
                "cav_loss":     "(total_cav - run_cav) / total_cav  (Sheet1 AK6)",
                "quality_loss": "rej / (rej + good)                 (Sheet1 X6)",
                "speed_loss":   "residual: 1 - oee_running/(avail*cav*quality)  (Sheet1 D7)",
                "downtime_pct": "each category hrs / total shift hrs  (Sheet2)",
                "not_run_pct":  "count / total machine-slots          (Sheet2)",
            },
        },
        "date_range": date_range,
        "aggregate":  aggregate,
        "daily":      daily_data,
        "machines":   machines,
        "records":    all_records,
        "not_run":    all_not_run,
        "all_slots":  all_slots,
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, default=str)

    log.info(f"\nDone. Written to: {out_path.resolve()}")
    log.info(f"  Dates:   {date_range['from']}  to  {date_range['to']}  ({date_range['days']} day(s))")
    log.info(f"  Records: {len(all_records)} machine-shift rows across {len(machines)} machines")

    if failed_files:
        log.warning(f"  Skipped {len(failed_files)} file(s): {failed_files}")

    # Console waterfall summary
    d = aggregate
    print("\n" + "="*55)
    print(f"  WATERFALL  ({date_range['from']}  to  {date_range['to']})")
    print("="*55)
    sr = d["shift_run_pct"]
    av = 1 - d["avail_loss"]
    cv = 1 - d["cav_loss"]
    sp = 1 - d["speed_loss"]
    print(f"  Theoretical max          100.00%")
    print(f"  x Shift run %          {sr*100:8.2f}%   loss {(1-sr)*100:.2f}%  ({d['not_ran_slots']}/{d['total_slots']} slots idle)")
    print(f"  x Availability         {sr*av*100:8.2f}%   loss {d['avail_loss']*100:.2f}%  (downtime)")
    print(f"  x Cavity perf          {sr*av*cv*100:8.2f}%   loss {d['cav_loss']*100:.2f}%  (fewer cavities)")
    print(f"  x Speed (residual)     {sr*av*cv*sp*100:8.2f}%   loss {d['speed_loss']*100:.2f}%")
    print(f"  x Quality              {d['overall_oee']*100:8.2f}%   loss {d['quality_loss']*100:.2f}%  (rejection)")
    print(f"\n  OVERALL OEE            {d['overall_oee']*100:8.2f}%")
    print(f"  OEE (running only)     {d['oee_running']*100:8.2f}%")
    print(f"  Good output       {d['total_actual']:>12,} pcs")
    print(f"  Target            {d['total_target']:>12,} pcs")
    print("="*55 + "\n")


if __name__ == "__main__":
    main()
